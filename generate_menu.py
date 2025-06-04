from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.simplefilter("ignore", UserWarning)

# Charger le fichier Excel
fichier_excel = "Analyse_Risques_Biens.xlsx"
wb = load_workbook(filename=fichier_excel, read_only=True)

# Feuilles nécessaires
ws_criteres = wb['RISQUES_BIENS']
ws_osi_layers = wb['RISQUES_ATTAQUES']
ws_attack_types = wb['ATTAQUES_OUTILS']

# Critères de sécurité fixes
criteres = ["Confidentialité", "Intégrité", "Disponibilité"]

# Charger le DataFrame pour les couches OSI
df_osi_layers = pd.read_excel(fichier_excel, sheet_name="RISQUES_ATTAQUES")
osi_row = df_osi_layers.iloc[0]

if pd.isna(osi_row.iloc[0]):
    osi_layers = ["Couche 1"] + [
        f"Couche {str(int(float(val)))}"
        for i, val in enumerate(osi_row[1:].values) if pd.notna(val)
    ]
else:
    osi_layers = [
        f"{i+1}-Couche {str(int(float(val)))}"
        for i, val in enumerate(osi_row.values) if pd.notna(val)
    ]

attaques = [row[0] for row in ws_attack_types.iter_rows(min_row=2, values_only=True) if row[0]]
attaques_formatees = [f"{attaque}" for i, attaque in enumerate(attaques)]

with open("menu_interactif.sh", "w") as f:
    f.write("""#!/bin/bash

# --- MENU GÉNÉRÉ AUTOMATIQUEMENT ---

normalize_string() {
    echo "$1" | iconv -f UTF-8 -t ASCII//TRANSLIT
}

""")
    f.write("criteres=(" + " ".join(f'"{c}"' for c in criteres) + ")\n")
    f.write("osi_layers=(" + " ".join(f'"{c}"' for c in osi_layers) + ")\n")
    f.write("attaques=(" + " ".join(f'"{a}"' for a in attaques_formatees) + ")\n")

    f.write("""
search_terms=()
logic="OU"  # Logique par défaut à OU

ajouter_critere() {
    echo
    fichiers=""
    if [ "$logic" == "ET" ]; then
        fichiers=$(find /home/$USER/cheatsheet -type f)
        for term in "${search_terms[@]}"; do
            fichiers=$(grep -il "$term" $fichiers 2>/dev/null)
        done
    else
        fichiers=$(grep -il "${search_terms[@]}" /home/$USER/cheatsheet/* 2>/dev/null)
    fi

    fichiers_count=$(echo "$fichiers" | grep -c .)
    echo "Pour le moment $fichiers_count réponse(s) trouvée(s)"
    
    if [ "$fichiers_count" -eq 0 ]; then
        echo "Critère trop restrictif. Pour le moment je n’ai rien en base."
        exit 0
    fi

    read -p "Voulez-vous completer par un autre critere ? (O/N) : " continuer
    continuer=$(echo "$continuer" | tr '[:lower:]' '[:upper:]')
    case $continuer in
        O) 
            read -p "Souhaitez-vous appliquer une logique 'ET' (recherche stricte) ou 'OU' (recherche large) ? (ET/OU) : " logique
            logique=$(echo "$logique" | tr '[:lower:]' '[:upper:]')
            if [[ "$logique" == "ET" || "$logique" == "OU" ]]; then
                logic="$logique"
            else
                echo "DONNÉE INCORRECTE, logique par défaut appliquée (OU)."
                logic="OU"
            fi
            main_menu ;;
        N) lancer_recherche ;;
        Q) echo "Sortie." ; exit 0 ;;
        *) echo "DONNÉE INCORRECTE" ; ajouter_critere ;;
    esac
}

lancer_recherche() {
    echo
    if [ "$logic" == "ET" ]; then
        fichiers=$(find /home/$USER/cheatsheet -type f)
        for term in "${search_terms[@]}"; do
            fichiers=$(grep -il "$term" $fichiers 2>/dev/null)
        done
    else
        fichiers=$(grep -il "${search_terms[@]}" /home/$USER/cheatsheet/* 2>/dev/null)
    fi

    if [ -z "$fichiers" ]; then
        echo "Critère trop restrictif. Pour le moment je n’ai rien en base."
    else
        echo "Résultat de la recherche :"
        echo "$fichiers" | xargs -I {} echo "{}"
    fi
    exit 0
}

main_menu() {
    echo
    echo "Que souhaitez-vous faire ?"
    echo "1 - Recherche sur les critères de sécurité"
    echo "2 - Recherche selon les couches du modèle OSI"
    echo "3 - Liste des types d'attaque"
    echo "Q - Quitter"
    read -p "Votre choix (1/2/3/Q) : " choix
    choix=$(echo "$choix" | tr '[:lower:]' '[:upper:]')
    case $choix in
        1)
            echo "Critères de sécurité :"
            for i in "${!criteres[@]}"; do echo "$((i+1)) - ${criteres[i]}"; done
            read -p "Choisissez un critère (1/2/3 ou Q pour quitter) : " critere_choix
            critere_choix=$(echo "$critere_choix" | tr '[:lower:]' '[:upper:]')
            [[ "$critere_choix" == "Q" ]] && echo "Sortie." && exit 0
            if [[ "$critere_choix" =~ ^[1-3]$ ]]; then
                selection="${criteres[$((critere_choix-1))]}"
                motcle=$(echo "$selection" | cut -d'-' -f2)
                motcle_normalize=$(normalize_string "$motcle")
                search_terms+=("$motcle_normalize")
                ajouter_critere
            else
                echo "DONNÉE INCORRECTE"
                main_menu
            fi
            ;;
        2)
            echo "Couches du modèle OSI :"
            for i in "${!osi_layers[@]}"; do echo "$((i+1)) - ${osi_layers[i]}"; done
            read -p "Choisissez une couche (numéro ou Q pour quitter) : " osi_choix
            osi_choix=$(echo "$osi_choix" | tr '[:lower:]' '[:upper:]')
            [[ "$osi_choix" == "Q" ]] && echo "Sortie." && exit 0
            if [[ "$osi_choix" =~ ^[0-9]+$ ]] && [ "$osi_choix" -ge 1 ] && [ "$osi_choix" -le ${#osi_layers[@]} ]; then
                selection="${osi_layers[$((osi_choix-1))]}"
                motcle=$(echo "$selection" | cut -d'-' -f2-)
                motcle_normalize=$(normalize_string "$motcle")
                search_terms+=("$motcle_normalize")
                ajouter_critere
            else
                echo "DONNÉE INCORRECTE"
                main_menu
            fi
            ;;
        3)
            echo "Types d'attaque disponibles :"
            for i in "${!attaques[@]}"; do echo "$((i+1)) - ${attaques[i]}"; done
            read -p "Choisissez un type d'attaque (numéro ou Q pour quitter) : " attaque_choix
            attaque_choix=$(echo "$attaque_choix" | tr '[:lower:]' '[:upper:]')
            [[ "$attaque_choix" == "Q" ]] && echo "Sortie." && exit 0
            if [[ "$attaque_choix" =~ ^[0-9]+$ ]] && [ "$attaque_choix" -ge 1 ] && [ "$attaque_choix" -le ${#attaques[@]} ]; then
                selection="${attaques[$((attaque_choix-1))]}"
                motcle=$(echo "$selection" | cut -d'-' -f2-)
                motcle_normalize=$(normalize_string "$motcle")
                search_terms+=("$motcle_normalize")
                ajouter_critere
            else
                echo "DONNÉE INCORRECTE"
                main_menu
            fi
            ;;
        Q) echo "Sortie." ; exit 0 ;;
        *) echo "DONNÉE INCORRECTE" ; main_menu ;;
    esac
}

main_menu
""")
